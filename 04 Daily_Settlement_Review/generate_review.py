import openpyxl
import os
import shutil

BASE = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(BASE, 'assets')
OUTPUT = os.path.join(BASE, 'output')
os.makedirs(OUTPUT, exist_ok=True)

TEMPLATE_PATH = os.path.join(ASSETS, '输出模版-0504-日结算收益复盘.xlsx')
DATES = ['0511', '0512', '0513', '0514', '0515', '0516']


def _is_merged(cell):
    """Check if a cell is a non-writable MergedCell."""
    return type(cell).__name__ == 'MergedCell'


def convert_to_numeric(ws):
    """Convert string values that look like numbers to float/int.
    Skips formula cells and MergedCells."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if _is_merged(cell) or cell.value is None:
                continue
            if not isinstance(cell.value, str):
                continue
            if str(cell.value).startswith('='):
                continue
            s = str(cell.value).strip()
            try:
                v = float(s)
                if v == int(v) and '.' not in s and 'e' not in s.lower():
                    cell.value = int(v)
                else:
                    cell.value = v
            except ValueError:
                pass


def copy_sheet_data(src_ws, dst_ws):
    """Copy non-None values from src to dst, preserving dst formulas.
    Skips MergedCells and formula cells in dst."""
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, max_col=src_ws.max_column):
        for src_cell in row:
            if src_cell.value is None:
                continue
            dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column)
            if _is_merged(dst_cell):
                continue
            # Never overwrite formulas in the target
            if isinstance(dst_cell.value, str) and str(dst_cell.value).startswith('='):
                continue
            dst_cell.value = src_cell.value
            if src_cell.number_format and src_cell.number_format != 'General':
                dst_cell.number_format = src_cell.number_format


def compute_J4(rt_review_path):
    """Compute capacity allocation coefficient (J4) from RT review data.
    Weighted average of monthly coefficients during charging periods."""
    wb = openpyxl.load_workbook(rt_review_path, data_only=True)
    cap_ws = wb['容量分摊系数']
    price_ws = wb['报价及预中标']

    cap_numer = 0.0
    cap_denom = 0.0

    for r in range(8, cap_ws.max_row + 1):
        price_row = r - 6  # cap row 8 → price row 2
        coeff = cap_ws.cell(row=r, column=10).value  # J column = 5月
        n_val = price_ws.cell(row=price_row, column=14).value  # N column

        if coeff is None or n_val is None:
            continue
        try:
            coeff = float(coeff)
            n_val = float(n_val)
        except (ValueError, TypeError):
            continue

        if n_val <= 0:  # charging period
            charge_mwh = n_val / 4
            cap_numer += coeff * charge_mwh
            cap_denom += charge_mwh

    wb.close()
    return cap_numer / cap_denom if cap_denom != 0 else 0


def generate_review(date_mmdd):
    """Generate daily settlement review from template, replacing only data cells.
    Preserves all template formulas, formats, and merged cells."""
    day = date_mmdd[2:4]
    date_iso = f'2026-05-{day}'

    charge_stmt = os.path.join(ASSETS, f'6052-{date_iso}德州润津储能科技有限公司结算单-充电.xlsx')
    discharge_stmt = os.path.join(ASSETS, f'6052-{date_iso}德州润津储能科技有限公司结算单-放电.xlsx')
    rt_review = os.path.join(ASSETS, f'{date_mmdd}-实时机组组合收益复盘.xlsx')
    out_path = os.path.join(OUTPUT, f'{date_mmdd}-日结算收益复盘.xlsx')

    for f in [charge_stmt, discharge_stmt, rt_review]:
        if not os.path.exists(f):
            print(f'  SKIP {date_mmdd}: Missing {os.path.basename(f)}')
            return False

    # 1. Copy template → output (preserves all formulas and formatting)
    shutil.copy2(TEMPLATE_PATH, out_path)

    # 2. Load source files
    wb_charge = openpyxl.load_workbook(charge_stmt, data_only=True)
    wb_discharge = openpyxl.load_workbook(discharge_stmt, data_only=True)
    wb_rt = openpyxl.load_workbook(rt_review, data_only=True)
    wb_out = openpyxl.load_workbook(out_path)

    # 3. 充电日清算费用 ← charge settlement 日清算数据 (full copy, no formulas in source)
    copy_sheet_data(wb_charge['日清算数据'], wb_out['充电日清算费用'])
    convert_to_numeric(wb_out['充电日清算费用'])

    # 4. 放电日清算费用 ← discharge settlement 日清算费用 (full copy, no formulas in source)
    copy_sheet_data(wb_discharge['日清算费用'], wb_out['放电日清算费用'])
    convert_to_numeric(wb_out['放电日清算费用'])

    # 5. 报价及预中标: ONLY write data columns H, J, K, N (rows 2-97)
    #    Template formulas in I, L, M and summary rows B25-D30 are preserved.
    rt_price = wb_rt['报价及预中标']
    out_price = wb_out['报价及预中标']
    for i in range(96):
        row = i + 2
        for col in [8, 10, 11, 14]:  # H=time, J=电价, K=电价, N=充放电
            src_val = rt_price.cell(row=row, column=col).value
            if src_val is not None:
                dst_cell = out_price.cell(row=row, column=col)
                if not _is_merged(dst_cell):
                    if col == 8:
                        dst_cell.value = str(src_val)
                    else:
                        try:
                            dst_cell.value = float(src_val)
                        except (ValueError, TypeError):
                            dst_cell.value = src_val

    # 6. 充放测算: ONLY update J4 and I8-I14 parameters. All row 4 formulas preserved.
    out_cf = wb_out['充放测算']
    rt_cf = wb_rt['充放测算']

    j4_val = compute_J4(rt_review)
    j4_cell = out_cf.cell(row=4, column=10)
    if not _is_merged(j4_cell):
        j4_cell.value = round(j4_val, 10)

    for r in [8, 9, 12, 13, 14]:
        src_val = rt_cf.cell(row=r, column=9).value
        if src_val is not None:
            dst_cell = out_cf.cell(row=r, column=9)
            if not _is_merged(dst_cell):
                try:
                    dst_cell.value = float(src_val)
                except (ValueError, TypeError):
                    dst_cell.value = src_val

    # 7. 容量分摊系数: NOT touched — template formulas reference 报价及预中标 columns,
    #    which now contain the date-specific data. Let Excel compute them.

    wb_out.save(out_path)
    for wb in [wb_charge, wb_discharge, wb_rt, wb_out]:
        wb.close()
    print(f'[OK] {os.path.basename(out_path)}')
    return True


if __name__ == '__main__':
    print(f'Template: {os.path.basename(TEMPLATE_PATH)}')
    print(f'Generating daily settlement reviews for dates: {", ".join(DATES)}\n')
    for d in DATES:
        generate_review(d)
    print('\nDone.')
