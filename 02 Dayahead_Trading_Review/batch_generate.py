"""Batch generate day-ahead review files for multiple dates."""
import sys
from pathlib import Path
import pandas as pd
import openpyxl

ROOT = Path(__file__).parent
TPL_PATH = ROOT / "assets" / "0505-日前机组组合收益复盘.xlsx"

# source file name patterns: base name for 0508, numbered suffixes for 0509-0520
SRC_FILES = {
    "0508": "0508-发电侧日前交易结果查询.xls",
}
for d in range(9, 21):
    mmdd = f"05{d:02d}"
    idx = d - 8  # 0509 → (1), 0520 → (12)
    SRC_FILES[mmdd] = f"{mmdd}-发电侧日前交易结果查询 ({idx}).xls"


def generate_one(mmd: str, src_name: str) -> None:
    src = ROOT / "assets" / src_name
    out = ROOT / "output" / f"{mmd}-日前机组组合收益复盘.xlsx"

    df = pd.read_excel(src, header=None)
    wb = openpyxl.load_workbook(TPL_PATH)
    ws = wb["报价及预中标"]

    for i in range(96):
        row = i + 2
        power = float(df.iloc[i + 1, 1])
        price = float(df.iloc[i + 1, 3])
        ws.cell(row=row, column=10, value=price)
        ws.cell(row=row, column=11, value=price)
        ws.cell(row=row, column=14, value=power)

    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"[OK] {out}")


def main():
    for mmd in sorted(SRC_FILES):
        generate_one(mmd, SRC_FILES[mmd])


if __name__ == "__main__":
    main()
