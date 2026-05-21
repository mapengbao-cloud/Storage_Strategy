import sys
from pathlib import Path
import pandas as pd
import openpyxl


def generate(mmd: str, src: str | None = None) -> None:
    """Generate day-ahead review file for a given MM-DD date string.

    Args:
        mmd: MM-DD date string (e.g. '0504').
        src: Optional path to source xls. Defaults to assets/发电侧日前交易结果查询.xls.
    """
    root = Path(__file__).parent
    src_path = Path(src) if src else root / "assets" / f"{mmd}-发电侧日前交易结果查询.xls"
    tpl_path = root / "assets" / f"{mmd}-日前机组组合收益复盘.xlsx"
    out_path = root / "output" / f"{mmd}-日前机组组合收益复盘.xlsx"

    df = pd.read_excel(src_path, header=None)
    wb = openpyxl.load_workbook(tpl_path)

    ws = wb["报价及预中标"]

    for i in range(96):
        row = i + 2  # template rows 2-97
        power = float(df.iloc[i + 1, 1])   # col 1: 出力
        price = float(df.iloc[i + 1, 3])   # col 3: 电价

        ws.cell(row=row, column=10, value=price)   # J: 统一结算日前
        ws.cell(row=row, column=11, value=price)   # K: 节点日前电价
        ws.cell(row=row, column=14, value=power)   # N: 充放电曲线

    out_path.parent.mkdir(exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    mmd = sys.argv[1]
    src = sys.argv[2] if len(sys.argv) > 2 else None
    generate(mmd, src)
