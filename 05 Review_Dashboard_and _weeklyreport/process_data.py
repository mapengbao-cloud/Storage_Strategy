import openpyxl
import os
import re
import shutil
import argparse
from datetime import datetime
from collections import defaultdict

# Paths
ASSETS_DIR = 'assets'
OUTPUT_DIR = 'output'
TARGET_FILE = '山东夏津储能收益统计表.xlsx'

# Target column mapping for each source type
# Source columns A(1)-Q(17) -> target columns for each section
# 日前机组组合收益复盘: B(2)-R(18)
# 实时机组组合收益复盘: S(19)-AI(35)
# 日结算收益复盘: AJ(36)-AX(50), skip AY(51), AZ(52)-BA(53)

def source_to_target_mapping(start_col):
    """Map source columns A-Q (1-17) to target columns starting at start_col.
    Returns dict: source_col_index -> target_col_index"""
    mapping = {}
    for i in range(17):  # source columns A-Q (0-16)
        mapping[i + 1] = start_col + i  # source col 1 (A) -> target start_col
    return mapping

# 日前 starts at target column B (2)
day_ahead_mapping = source_to_target_mapping(2)
# 实时 starts at target column S (19)
real_time_mapping = source_to_target_mapping(19)

# 日结算 has an extra column AY (51) between AX (50) and AZ (52).
# Source A-O (1-15) -> Target AJ-AX (36-50) linear
# Source P (16) -> Target AZ (52), Source Q (17) -> Target BA (53)
settlement_mapping = {}
for i in range(1, 16):
    settlement_mapping[i] = 36 + (i - 1)  # A-O -> AJ-AX
settlement_mapping[16] = 52  # Source P -> AZ
settlement_mapping[17] = 53  # Source Q -> BA

# For 日结算, column AY (51) gets O6 value
SETTLEMENT_AY_COL = 51

# Columns that use '0.00%' format (综合效率)
PCT_COLUMNS = {17, 34, 52}  # Q, AH, AZ in target

def clean_value(val, target_col=None):
    """Round numeric values to 2 decimal places. 综合效率 (PCT_COLUMNS) keeps
    original precision since it's stored as a percentage decimal."""
    if isinstance(val, float):
        if target_col is not None and target_col in PCT_COLUMNS:
            return val  # preserve full precision for 综合效率
        return round(val, 2)
    return val

def parse_date_from_filename(filename):
    """Extract MMDD date from filename and return (month, day)."""
    match = re.match(r'^(\d{2})(\d{2})', filename)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None

def find_date_in_target(ws, month, day):
    """Find the row in target sheet matching month/day.
    Handles both datetime objects and Excel serial numbers."""
    excel_epoch = datetime(1899, 12, 30)
    for year in [2026, 2025]:
        try:
            target_serial = (datetime(year, month, day) - excel_epoch).days
        except ValueError:
            continue
        for row in range(4, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=1).value
            if isinstance(cell_val, datetime):
                if (cell_val.year == year and
                    cell_val.month == month and
                    cell_val.day == day):
                    return row
            elif isinstance(cell_val, (int, float)) and cell_val > 40000:
                if int(cell_val) == target_serial:
                    return row
    return None

def compute_summary_values(source_path):
    """Compute the 17 summary values (row 4 equivalent) and O6 value
    from raw data in the source workbook's 报价及预中标 sheet.
    Used for 日前 and 实时 files where 充放测算 row 4 formulas
    reference 报价及预中标 internally.
    Returns (list_of_17_values, o6_value)."""
    wb = openpyxl.load_workbook(source_path, data_only=False)

    # Find sheets by name pattern (index varies by file type)
    sn0 = wb.sheetnames[0]  # 充放测算 is always first
    sn_price = next(s for s in wb.sheetnames if '报价' in s and '预中标' in s)
    sn_cap = next(s for s in wb.sheetnames if '容量分摊' in s)

    ws0 = wb[sn0]
    ws1 = wb[sn_price]
    ws2 = wb[sn_cap]

    # Read raw data: col J (电价), col N (充放电需求)
    charge_vol = 0.0
    discharge_vol = 0.0
    charge_price_sum = 0.0
    discharge_price_sum = 0.0

    for row in range(2, ws1.max_row + 1):
        j = ws1.cell(row=row, column=10).value
        n = ws1.cell(row=row, column=14).value
        if j is None or n is None:
            continue
        try:
            j = float(j)
            n = float(n)
        except (ValueError, TypeError):
            continue
        if n <= 0:
            charge_vol += n
            charge_price_sum += j * n
        if n >= 0:
            discharge_vol += n
            discharge_price_sum += j * n

    B = charge_vol / 4          # 充电量
    M_val = discharge_vol / 4   # 放电量
    A = charge_price_sum / charge_vol if charge_vol != 0 else 0   # 充电价
    L = discharge_price_sum / discharge_vol if discharge_vol != 0 else 0  # 放电价

    # Parameters from 充放测算
    I8 = float(ws0.cell(row=8, column=9).value or 0)
    I9 = float(ws0.cell(row=9, column=9).value or 0)
    I12 = float(ws0.cell(row=12, column=9).value or 0)
    I13 = float(ws0.cell(row=13, column=9).value or 0)
    I14 = float(ws0.cell(row=14, column=9).value or 0)

    # Capacity allocation coefficient from 容量分摊系数
    month_col = 10  # 5月
    cap_numer = 0.0
    cap_denom = 0.0
    for row in range(8, ws2.max_row + 1):
        coeff = ws2.cell(row=row, column=month_col).value
        n_val = ws1.cell(row=row, column=14).value
        if coeff is None or n_val is None:
            continue
        try:
            coeff = float(coeff)
            n_val = float(n_val)
        except (ValueError, TypeError):
            continue
        if n_val <= 0:
            charge_mwh = n_val / 4
            cap_numer += coeff * charge_mwh
            cap_denom += charge_mwh

    J_val = cap_numer / cap_denom if cap_denom != 0 else 0
    P = -M_val / B if B != 0 else 0

    C = A * B
    D = B * I12
    E = B * I13
    F = B * (1 - P) * I8
    G = B * (1 - P) * I9
    H = B * I14
    I_val = B * J_val * 70.5
    K = C + D + E + F + G + H + I_val
    N_val = L * M_val
    O_val = N_val + K
    Q = L - A

    o6 = N_val + C

    wb.close()
    return [A, B, C, D, E, F, G, H, I_val, J_val, K, L, M_val, N_val, O_val, P, Q], o6


def compute_settlement_values(source_path):
    """Compute the 17 summary values and O6 from a 日结算收益复盘 file.

    The 日结算 file's 充放测算 row 4 formulas reference settlement sheets
    (充电日清算费用 / 放电日清算费用), NOT 报价及预中标. This function
    reads the base values from those settlement sheets and computes
    the dependent values (D-Q) using the same formula logic as 充放测算 row 4.

    充放测算 row 4 formula sources:
      A4 = 充电日清算费用!AA29    (col 27)
      B4 = -充电日清算费用!Z29     (col 26) → negated
      C4 = -充电日清算费用!AB29    (col 28) → negated
      L4 = 放电日清算费用!P101     (col 16)
      M4 = 放电日清算费用!AJ101    (col 36)
      N4 = 放电日清算费用!AK101    (col 37)
      O6 = N4 + C4

    Returns (list_of_17_values, o6_value)."""
    wb = openpyxl.load_workbook(source_path, data_only=True)

    ws_cf = wb['充放测算']
    ws_charge = wb['充电日清算费用']
    ws_discharge = wb['放电日清算费用']

    # Base values from settlement sheets
    A = float(ws_charge.cell(row=29, column=27).value or 0)      # AA29 充电价
    B = -float(ws_charge.cell(row=29, column=26).value or 0)     # Z29 → 充电量
    C = -float(ws_charge.cell(row=29, column=28).value or 0)     # AB29 → 充电收入
    L = float(ws_discharge.cell(row=101, column=16).value or 0)  # P101 放电价
    M_val = float(ws_discharge.cell(row=101, column=36).value or 0)  # AJ101 放电量
    N_val = float(ws_discharge.cell(row=101, column=37).value or 0)  # AK101 放电收入

    # Parameters from 充放测算
    I8 = float(ws_cf.cell(row=8, column=9).value or 0)
    I9 = float(ws_cf.cell(row=9, column=9).value or 0)
    I12 = float(ws_cf.cell(row=12, column=9).value or 0)
    I13 = float(ws_cf.cell(row=13, column=9).value or 0)
    I14 = float(ws_cf.cell(row=14, column=9).value or 0)
    J_val = float(ws_cf.cell(row=4, column=10).value or 0)  # 容量分摊系数

    # Compute dependent values using same formulas as 充放测算 row 4
    P = -M_val / B if B != 0 else 0
    D = B * I12
    E_val = B * I13
    F = B * (1 - P) * I8
    G = B * (1 - P) * I9
    H = B * I14
    I_val = B * J_val * 70.5
    K = C + D + E_val + F + G + H + I_val
    O_val = N_val + K
    Q = L - A
    o6 = N_val + C

    wb.close()
    return [A, B, C, D, E_val, F, G, H, I_val, J_val, K, L, M_val, N_val, O_val, P, Q], o6


def parse_range_arg(arg):
    """Parse 'MMDD-MMDD' range string, return (month, start_day, end_day)."""
    m = re.match(r'^(\d{2})(\d{2})-(\d{2})(\d{2})$', arg)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(4)))
    return None

def in_range(month, day, range_tuple):
    """Check if (month, day) is within the range (inclusive). None = no filter."""
    if range_tuple is None:
        return True
    rm, rsd, red = range_tuple
    if month != rm:
        return False
    return rsd <= day <= red


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Process energy storage revenue review data.')
    parser.add_argument('--day-ahead', type=str, default=None,
                        help='Date range for 日前 data, e.g. 0508-0520')
    parser.add_argument('--real-time', type=str, default=None,
                        help='Date range for 实时 data, e.g. 0511-0519')
    parser.add_argument('--settlement', type=str, default=None,
                        help='Date range for 日结算 data, e.g. 0511-0516')
    args = parser.parse_args()

    day_ahead_filter = parse_range_arg(args.day_ahead) if args.day_ahead else None
    real_time_filter = parse_range_arg(args.real_time) if args.real_time else None
    settlement_filter = parse_range_arg(args.settlement) if args.settlement else None

    if day_ahead_filter:
        print(f"日前 filter: {args.day_ahead}")
    if real_time_filter:
        print(f"实时 filter: {args.real_time}")
    if settlement_filter:
        print(f"日结算 filter: {args.settlement}")

    # Group source files by date
    source_files = defaultdict(dict)

    for fname in os.listdir(ASSETS_DIR):
        if not fname.endswith('.xlsx') or fname.startswith('~$'):
            continue
        if fname == TARGET_FILE:
            continue

        date_info = parse_date_from_filename(fname)
        if date_info is None:
            continue

        month, day = date_info

        if '日前机组组合收益复盘' in fname:
            source_files[(month, day)]['day_ahead'] = fname
        elif '实时机组组合收益复盘' in fname:
            source_files[(month, day)]['real_time'] = fname
        elif '日结算收益复盘' in fname:
            source_files[(month, day)]['settlement'] = fname

    print("Found source files by date:")
    for (m, d), files in sorted(source_files.items()):
        print(f"  {m:02d}/{d:02d}: {files}")

    # Use previous output as base (incremental), fall back to template
    src_tpl = os.path.join(ASSETS_DIR, TARGET_FILE)
    prev_out = os.path.join(OUTPUT_DIR, TARGET_FILE)
    out_path = prev_out

    if os.path.exists(prev_out):
        # Previous output exists: use it as base to preserve existing data
        base_src = prev_out
        print(f"\nUsing previous output as base: {prev_out}")
    else:
        base_src = src_tpl
        print(f"\nNo previous output, using template: {src_tpl}")

    try:
        shutil.copy2(base_src, out_path)
    except PermissionError:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        name, ext = os.path.splitext(TARGET_FILE)
        out_path = os.path.join(OUTPUT_DIR, f'{name}_{ts}{ext}')
        shutil.copy2(base_src, out_path)
    print(f"Output: {out_path}")

    # Open target workbook for editing
    wb = openpyxl.load_workbook(out_path)
    ws = wb['润津']

    processed_count = 0

    for (month, day), files in sorted(source_files.items()):
        target_row = find_date_in_target(ws, month, day)
        if target_row is None:
            print(f"  WARNING: Date {month:02d}/{day:02d} "
                  f"not found in target, skipping")
            continue

        print(f"\nProcessing date {month:02d}/{day:02d} "
              f"-> target row {target_row}")

        # Process 日前机组组合收益复盘
        if 'day_ahead' in files and in_range(month, day, day_ahead_filter):
            fname = files['day_ahead']
            print(f"  - 日前: {fname}")
            src_fpath = os.path.join(ASSETS_DIR, fname)
            src_values, _ = compute_summary_values(src_fpath)
            for src_col, val in enumerate(src_values, 1):
                target_col = day_ahead_mapping[src_col]
                cell = ws.cell(row=target_row, column=target_col)
                cell.value = clean_value(val, target_col)

        # Process 实时机组组合收益复盘
        if 'real_time' in files and in_range(month, day, real_time_filter):
            fname = files['real_time']
            print(f"  - 实时: {fname}")
            src_fpath = os.path.join(ASSETS_DIR, fname)
            src_values, _ = compute_summary_values(src_fpath)
            for src_col, val in enumerate(src_values, 1):
                target_col = real_time_mapping[src_col]
                cell = ws.cell(row=target_row, column=target_col)
                cell.value = clean_value(val, target_col)

        # Process 日结算收益复盘
        if 'settlement' in files and in_range(month, day, settlement_filter):
            fname = files['settlement']
            print(f"  - 日结算: {fname}")
            src_fpath = os.path.join(ASSETS_DIR, fname)
            src_values, o6_val = compute_settlement_values(src_fpath)
            for src_col, val in enumerate(src_values, 1):
                target_col = settlement_mapping[src_col]
                cell = ws.cell(row=target_row, column=target_col)
                cell.value = clean_value(val, target_col)

            # O6 -> AY column
            print(f"    O6 value: {o6_val} -> AY{target_row}")
            ay_cell = ws.cell(row=target_row, column=SETTLEMENT_AY_COL)
            ay_cell.value = clean_value(o6_val)

        processed_count += 1

    # Save the modified workbook
    wb.save(out_path)
    wb.close()

    print(f"\nDone! Processed {processed_count} dates. "
          f"Output saved to: {out_path}")
